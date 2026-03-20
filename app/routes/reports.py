import io
from datetime import datetime
from flask import Blueprint, render_template, request, send_file, jsonify, flash, redirect, url_for
from flask_login import login_required, current_user
from app import db
from app.models import PowerOfAttorney, Client
from app.utils import log_activity

reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


def _build_query():
    """بناء استعلام مشترك بين المعاينة والتصدير."""
    q = PowerOfAttorney.query.filter_by(is_deleted=False)

    # اسم الموكل — بحث جزئي في جدول clients
    keyword = request.args.get("q", "").strip()
    if keyword:
        like = f"%{keyword}%"
        q = (q.join(Client, Client.poa_id == PowerOfAttorney.id, isouter=True)
              .filter(db.or_(Client.full_name.ilike(like), Client.national_id.ilike(like), Client.phone.ilike(like)))
              .distinct())

    # رقم التوكيل بالمكتب — بالضبط
    office_number = request.args.get("office_number", "").strip()
    if office_number:
        try:
            q = q.filter(PowerOfAttorney.office_number == int(office_number))
        except ValueError:
            pass

    # السنة
    year = request.args.get("year", "").strip()
    if year:
        try:
            q = q.filter(PowerOfAttorney.year == int(year))
        except ValueError:
            pass

    # الحالة
    status = request.args.get("status", "").strip()
    if status:
        q = q.filter(PowerOfAttorney.status == status)

    # حالة التفعيل
    activation = request.args.get("activation_status", "").strip()
    if activation:
        q = q.filter(PowerOfAttorney.activation_status == activation)

    # الحرف
    letter = request.args.get("letter", "").strip()
    if letter:
        q = q.filter(PowerOfAttorney.letter == letter)

    return q


@reports_bp.route("/")
@login_required
def index():
    total       = PowerOfAttorney.query.filter_by(is_deleted=False).count()
    active      = PowerOfAttorney.query.filter_by(is_deleted=False, status="ساري").count()
    expired     = PowerOfAttorney.query.filter_by(is_deleted=False, status="منتهي").count()
    active_only = PowerOfAttorney.query.filter_by(is_deleted=False, activation_status="مفعل").count()
    inactive    = PowerOfAttorney.query.filter_by(is_deleted=False, activation_status="غير مفعل").count()
    return render_template("reports/index.html",
                           total=total, active=active, expired=expired,
                           active_only=active_only, inactive=inactive)


@reports_bp.route("/preview")
@login_required
def preview():
    """API — إرجاع عدد النتائج قبل التصدير."""
    count = _build_query().count()
    return jsonify({"count": count})


@reports_bp.route("/export/excel")
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    poas = _build_query().order_by(PowerOfAttorney.office_number).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "التوكيلات"
    ws.sheet_view.rightToLeft = True

    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    header_fill = PatternFill("solid", start_color="1E3A5F")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "#", "رقم التوكيل بالمكتب", "اسم الموكل", "الرقم القومي",
        "رقم الهاتف", "العنوان", "رقم التوكيل", "مسمى التوكيل",
        "الحرف", "السنة", "مكتب التوثيق", "المحامي",
        "الحالة", "تاريخ انتهاء الصلاحية", "حالة التفعيل", "ملاحظات",
    ]
    col_widths = [5, 20, 28, 18, 16, 28, 16, 30, 8, 8, 22, 22, 10, 22, 14, 25]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 32

    fill_even   = PatternFill("solid", start_color="F0F4FA")
    normal_font = Font(name="Arial", size=10)

    row_num = 2
    for seq, poa in enumerate(poas, start=1):
        clients = list(poa.clients.all())
        if not clients:
            clients = [None]

        for c_idx, client in enumerate(clients):
            fill = fill_even if row_num % 2 == 0 else None
            row_data = [
                seq               if c_idx == 0 else "",
                poa.office_number if c_idx == 0 else "",
                client.full_name   if client else "—",
                (client.national_id or "") if client else "",
                (client.phone      or "") if client else "",
                (client.address    or "") if client else "",
                poa.poa_number        if c_idx == 0 else "",
                poa.poa_title         if c_idx == 0 else "",
                (poa.letter      or "") if c_idx == 0 else "",
                (poa.year        or "") if c_idx == 0 else "",
                (poa.notary_office or "") if c_idx == 0 else "",
                (poa.lawyer_name   or "") if c_idx == 0 else "",
                poa.status            if c_idx == 0 else "",
                (poa.expiry_date.strftime("%Y-%m-%d") if poa.expiry_date else "") if c_idx == 0 else "",
                poa.activation_status if c_idx == 0 else "",
                (poa.notes        or "") if c_idx == 0 else "",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell           = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font      = normal_font
                cell.alignment = center
                cell.border    = border
                if fill:
                    cell.fill = fill
            row_num += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_activity("EXPORT", "poa", None, f"تصدير {len(poas)} توكيل إلى Excel")
    db.session.commit()

    filename = f"توكيلات_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@reports_bp.route("/import", methods=["GET", "POST"])
@login_required
def import_excel():
    import pandas as pd

    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file or not file.filename.endswith((".xlsx", ".xls")):
            flash("يرجى رفع ملف Excel صحيح (.xlsx أو .xls)", "danger")
            return redirect(url_for("reports.import_excel"))

        try:
            df = pd.read_excel(file)
            df.columns = [c.strip() for c in df.columns]

            col_map = {
                "رقم التوكيل بالمكتب": "office_number",
                "موكل":                "client_name",
                "رقم التوكيل":         "poa_number",
                "مسمى التوكيل":        "poa_title",
                "الحرف":               "letter",
                "السنة":               "year",
                "مكتب التوثيق":        "notary_office",
                "الحالة":              "status_raw",
                "تاريخ انتهاء الصلاحية": "expiry_date",
                "حالة التفعيل":        "activation_status",
                "الرقم القومي":        "national_id",
                "رقم الهاتف":          "phone",
                "العنوان":             "address",
                "المحامي":             "lawyer_name",
            }
            df.rename(columns={k:v for k,v in col_map.items() if k in df.columns}, inplace=True)

            imported = skipped = errors = 0

            for _, row in df.iterrows():
                office_num = row.get("office_number")
                if pd.isna(office_num):
                    skipped += 1
                    continue
                try:
                    office_num = int(office_num)
                    if PowerOfAttorney.query.filter_by(office_number=office_num, is_deleted=False).first():
                        skipped += 1
                        continue

                    def _str(val):
                        return str(val).strip() if pd.notna(val) else None

                    status_raw = _str(row.get("status_raw")) or "ساري"
                    parts = [p.strip() for p in status_raw.split() if p.strip()]
                    status = next((p for p in reversed(parts) if p in ("ساري","منتهي","معلق")), "ساري")

                    expiry = None
                    if pd.notna(row.get("expiry_date")):
                        try: expiry = pd.to_datetime(row["expiry_date"]).date()
                        except: pass

                    activation = _str(row.get("activation_status")) or "مفعل"
                    if activation not in ("مفعل","غير مفعل"): activation = "مفعل"

                    poa = PowerOfAttorney(
                        office_number     = office_num,
                        poa_number        = str(int(row["poa_number"])) if pd.notna(row.get("poa_number")) else "0",
                        poa_title         = _str(row.get("poa_title")) or "توكيل عام",
                        letter            = _str(row.get("letter")),
                        year              = int(row["year"]) if pd.notna(row.get("year")) else None,
                        notary_office     = _str(row.get("notary_office")),
                        status            = status,
                        expiry_date       = expiry,
                        activation_status = activation,
                        lawyer_name       = _str(row.get("lawyer_name")),
                        created_by        = current_user.id,
                        updated_by        = current_user.id,
                    )
                    db.session.add(poa)
                    db.session.flush()
                    db.session.add(Client(
                        poa_id      = poa.id,
                        is_primary  = True,
                        full_name   = _str(row.get("client_name")) or "غير محدد",
                        national_id = _str(row.get("national_id")),
                        phone       = _str(row.get("phone")),
                        address     = _str(row.get("address")),
                    ))
                    imported += 1
                except Exception:
                    errors += 1
                    continue

            db.session.commit()
            log_activity("IMPORT", "poa", None, f"استيراد Excel: {imported} توكيل جديد، {skipped} مكرر، {errors} خطأ")
            db.session.commit()
            flash(f"✓ تم استيراد {imported} توكيل | مكرر: {skipped} | أخطاء: {errors}", "success")
        except Exception as e:
            flash(f"خطأ في قراءة الملف: {str(e)}", "danger")

        return redirect(url_for("reports.import_excel"))

    return render_template("reports/import.html")


@reports_bp.route("/stats")
@login_required
def stats():
    from sqlalchemy import func
    from datetime import date, timedelta

    # أكثر الموكلين
    top_clients = db.session.query(
        Client.full_name, func.count(Client.id).label("cnt")
    ).join(PowerOfAttorney).filter(
        PowerOfAttorney.is_deleted == False
    ).group_by(Client.full_name).order_by(func.count(Client.id).desc()).limit(10).all()

    # أكثر السنوات
    top_years = db.session.query(
        PowerOfAttorney.year, func.count(PowerOfAttorney.id).label("cnt")
    ).filter(
        PowerOfAttorney.is_deleted == False,
        PowerOfAttorney.year != None
    ).group_by(PowerOfAttorney.year).order_by(func.count(PowerOfAttorney.id).desc()).limit(10).all()

    # أكثر مكاتب التوثيق
    top_notary = db.session.query(
        PowerOfAttorney.notary_office, func.count(PowerOfAttorney.id).label("cnt")
    ).filter(
        PowerOfAttorney.is_deleted == False,
        PowerOfAttorney.notary_office != None
    ).group_by(PowerOfAttorney.notary_office).order_by(func.count(PowerOfAttorney.id).desc()).limit(10).all()

    # أكثر المحامين
    top_lawyers = db.session.query(
        PowerOfAttorney.lawyer_name, func.count(PowerOfAttorney.id).label("cnt")
    ).filter(
        PowerOfAttorney.is_deleted == False,
        PowerOfAttorney.lawyer_name != None
    ).group_by(PowerOfAttorney.lawyer_name).order_by(func.count(PowerOfAttorney.id).desc()).limit(10).all()

    # توكيلات تنتهي قريباً
    expiring_30 = PowerOfAttorney.query.filter(
        PowerOfAttorney.is_deleted == False,
        PowerOfAttorney.expiry_date != None,
        PowerOfAttorney.expiry_date <= date.today() + timedelta(days=30),
        PowerOfAttorney.expiry_date >= date.today(),
    ).count()

    total = PowerOfAttorney.query.filter_by(is_deleted=False).count()

    return render_template("reports/stats.html",
        top_clients=top_clients, top_years=top_years,
        top_notary=top_notary, top_lawyers=top_lawyers,
        expiring_30=expiring_30, total=total)